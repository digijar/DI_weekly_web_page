from fastapi import FastAPI, HTTPException, Body, Request, Response
import os
from google.cloud import bigquery
from json import JSONDecodeError
import json
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
from pydantic import BaseModel
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import datetime
from datetime import date

app = FastAPI()

# Enable CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = 'testing-bigquery-vertexai-service-account.json'
client = bigquery.Client()

@app.get("/download_mergermarket")
def download_MM():
    table_name = "MergerMarket"
    query = """
    SELECT
        Opportunity_ID,
        Date,
        `Value INR_m`,
        `Value Description`,
        Heading,
        Opportunity,
        Targets,
        `Lead type`,
        `Type of transaction`,
        `HS sector classification`,
        `Short BD`,
        Source,
        `Intelligence Type`,
        `Intelligence Grade`,
        `Intelligence Size`,
        `Stake Value`,
        `Dominant Sector`,
        Sectors,
        `Sub Sectors`,
        `Dominant Geography`,
        Geography,
        States,
        Topics,
        Bidders,
        Vendors,
        Issuers,
        Competitors,
        Others,
        Completed
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Opportunity_ID;
    """.format(table_name=table_name)

    # Run the BigQuery query
    job = client.query(query)

    # Fetch the query results
    result = job.result()

    # Extract data rows and header positions
    data_tuple = (
        [(row.Opportunity_ID, row.Date, row['Value INR_m'], row['Value Description'], row.Heading, row.Opportunity,
        row.Targets, row['Lead type'], row['Type of transaction'], row['HS sector classification'], row['Short BD'], row.Source,
        row['Intelligence Type'], row['Intelligence Grade'], row['Intelligence Size'], row['Stake Value'], row['Dominant Sector'],
        row.Sectors, row['Sub Sectors'], row['Dominant Geography'], row.Geography, row.States, row.Topics, row.Bidders,
        row.Vendors, row.Issuers, row.Competitors, row.Others, row.Completed)
        for row in result],
        {
            'Opportunity_ID': 0,
            'Date': 1,
            'Value INR_m': 2,
            'Value Description': 3,
            'Heading': 4,
            'Opportunity': 5,
            'Targets': 6,
            'Lead type': 7,
            'Type of transaction': 8,
            'HS sector classification': 9,
            'Short BD': 10,
            'Source': 11,
            'Intelligence Type': 12,
            'Intelligence Grade': 13,
            'Intelligence Size': 14,
            'Stake Value': 15,
            'Dominant Sector': 16,
            'Sectors': 17,
            'Sub Sectors': 18,
            'Dominant Geography': 19,
            'Geography': 20,
            'States': 21,
            'Topics': 22,
            'Bidders': 23,
            'Vendors': 24,
            'Issuers': 25,
            'Competitors': 26,
            'Others': 27,
            'Completed': 28
        }
    )

    # Extract row values and header positions
    row_values, header_positions = data_tuple

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Append the header row to the worksheet based on the header positions
    header_row = [None] * len(header_positions)
    for header, position in header_positions.items():
        header_row[position] = header
    ws.append(header_row)

    # Make the header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Append the data rows to the worksheet
    for row_data in row_values:
        # Check if row_data has at least 2 elements before converting date to string
        if len(row_data) >= 2 and isinstance(row_data[1], datetime.date):
            row_data = list(row_data)  # Convert tuple to list to allow modification
            row_data[1] = row_data[1].strftime('%Y-%m-%d')  # Convert date to string

        # Replace None values with an empty string
        row_data = ['' if value is None else value for value in row_data]

        # Replace '\r' and '\n' characters with line breaks in strings
        row_data = [value.replace('\r', '\n') if isinstance(value, str) else value for value in row_data]

        # Append the preprocessed row_data to the worksheet
        ws.append(row_data)

    # Create an in-memory file-like object to save the workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set response headers for file download
    headers = {
        'Content-Disposition': f'attachment; filename={"MergerMarket.xlsx"}',
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "POST, GET, OPTIONS",
    }
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return Response(
        content=output.read(),
        headers=headers,
        media_type=media_type
    )

@app.get('/get_mergermarket')
async def get_mergermarket():
    table_name = "MergerMarket"
    job = """
    SELECT
        Opportunity_ID,
        Date,
        `Value INR_m`,
        `Value Description`,
        Heading,
        Opportunity,
        Targets,
        `Lead type`,
        `Type of transaction`,
        `HS sector classification`,
        `Short BD`,
        Source,
        `Intelligence Type`,
        `Intelligence Grade`,
        `Intelligence Size`,
        `Stake Value`,
        `Dominant Sector`,
        Sectors,
        `Sub Sectors`,
        `Dominant Geography`,
        Geography,
        States,
        Topics,
        Bidders,
        Vendors,
        Issuers,
        Competitors,
        Others,
        Completed
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Opportunity_ID;
    """.format(table_name=table_name)

    result = client.query(job)

    return_json = []

    # Iterate through rows
    for row in result:
        row_dict = {}
        
        # Iterate through columns dynamically
        for column_name in row.keys():
            row_dict[column_name] = row[column_name]

        return_json.append(row_dict)

    return return_json


@app.post('/update_mergermarket')
async def update_mergermarket(request: Request):
    table_name = "MergerMarket"

    try:
        data = await request.json()

    except JSONDecodeError:
        return 'Invalid JSON data.'

    success = True  # Assume success initially

    try:
        for row_data in data:
            row_id = row_data.get("row_id")
            updated_column = row_data.get("column_name")
            updated_value = row_data.get("edited_value")

            # Construct the UPDATE query dynamically
            update_query = f"""
            UPDATE `testing-bigquery-vertexai.templates.{table_name}`
            SET `{updated_column}` = '{updated_value}'
            WHERE `Opportunity_ID` = {row_id}
            """
            client.query(update_query)
            
    except Exception as e:
        print(e)
        success = False
    
    return {"success": success}

class MergerMarketRow(BaseModel):
    Opportunity_ID: float
    Date: date
    Value_INR_m: str
    Value_Description: str
    Heading: str
    Opportunity: str
    Targets: str
    Lead_Type: str
    Type_of_transaction: str
    HS_sector_classification: str
    Short_BD: str
    Source: str
    Intelligence_Type: str
    Intelligence_Grade: str
    Intelligence_Size: str
    Stake_Value: str
    Dominant_Sector: str
    Sectors: str
    Sub_Sectors: str
    Dominant_Geography: str
    Geography: str
    States: str
    Topics: str
    Bidders: str
    Vendors: str
    Issuers: str
    Competitors: str
    Others: str
    Completed: str

@app.post('/add_mergermarket_row')
async def add_mergermarket_row(request: Request, row_data: MergerMarketRow):
    table_name = "MergerMarket"

    try:
        # Initialize the BigQuery client
        client = bigquery.Client()

        # Convert Date to ISO formatted string
        formatted_date = row_data.Date.isoformat()

        # Construct the INSERT query dynamically
        insert_query = f"""
        INSERT INTO `testing-bigquery-vertexai.templates.{table_name}`
        (Opportunity_ID, Date, `Value INR_m`, `Value Description`, Heading, Opportunity, Targets,
        `Lead type`, `Type of transaction`, `HS sector classification`, `Short BD`, Source,
        `Intelligence Type`, `Intelligence Grade`, `Intelligence Size`, `Stake Value`, `Dominant Sector`,
        Sectors, `Sub Sectors`, `Dominant Geography`, Geography, States, Topics, Bidders, Vendors,
        Issuers, Competitors, Others, Completed)
        VALUES 
        ({row_data.Opportunity_ID}, '{formatted_date}', '{row_data.Value_INR_m}', '{row_data.Value_Description}',
        '{row_data.Heading}', '{row_data.Opportunity}', '{row_data.Targets}', '{row_data.Lead_Type}',
        '{row_data.Type_of_transaction}', '{row_data.HS_sector_classification}', '{row_data.Short_BD}',
        '{row_data.Source}', '{row_data.Intelligence_Type}', '{row_data.Intelligence_Grade}',
        '{row_data.Intelligence_Size}', '{row_data.Stake_Value}', '{row_data.Dominant_Sector}',
        '{row_data.Sectors}', '{row_data.Sub_Sectors}', '{row_data.Dominant_Geography}', '{row_data.Geography}',
        '{row_data.States}', '{row_data.Topics}', '{row_data.Bidders}', '{row_data.Vendors}', '{row_data.Issuers}',
        '{row_data.Competitors}', '{row_data.Others}', '{row_data.Completed}')
        """

        # Run the query
        query_job = client.query(insert_query)

        # Wait for the query to complete (optional)
        query_job.result()

        return {"success": True}

    except Exception as e:
        print(e)
        return {"success": False}

@app.get("/download_marketscan")
def download_MS():
    table_name = "Market_Scan"
    query = """
    SELECT
        `Num`,
        `Picks`,
        `Captured in the week`,
        `Date`,
        `Lead type`,
        `Target HS industry classification`,
        `Dominant sector_as per MM`,
        `Sectors_as per MM`,
        `Sub Sectors_as per MM`,
        `HS Target region`,
        `Dominant country`,
        `Countries`,
        `Next Steps`,
        `Target`,
        `Target_Chinese`,
        `Deal intelligence`,
        `Short BD`,
        `Bidders`,
        `Sellers_or_Vendors`,
        `Type of transaction`,
        `Intelligence type_as per MM`,
        `Topics_as per MM`,
        `Value_USDm`,
        `Value description`,
        `Intelligence size`,
        `Intelligence size bucket`,
        `Stake value_percent`,
        `Held by ASPAC priority firm_Y_or_N`,
        `PE priority firm `,
        `Held since`,
        `Held for more than three years_Y_N_NA`,
        `KPMG credentials_PE_Company_Both_N`,
        `KPMG firm`,
        `Engagement partner`,
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Num;
    """.format(table_name=table_name)

    # Run the BigQuery query
    job = client.query(query)

    # Fetch the query results
    result = job.result()

    # Extract data rows and header positions
    data_tuple = (
        [
            (
                row.Num,
                row.Picks,
                row['Captured in the week'],
                row.Date,
                row['Lead type'],
                row['Target HS industry classification'],
                row['Dominant sector_as per MM'],
                row['Sectors_as per MM'],
                row['Sub Sectors_as per MM'],
                row['HS Target region'],
                row['Dominant country'],
                row.Countries,
                row['Next Steps'],
                row.Target,
                row['Target_Chinese'],
                row['Deal intelligence'],
                row['Short BD'],
                row.Bidders,
                row['Sellers_or_Vendors'],
                row['Type of transaction'],
                row['Intelligence type_as per MM'],
                row['Topics_as per MM'],
                row['Value_USDm'],
                row['Value description'],
                row['Intelligence size'],
                row['Intelligence size bucket'],
                row['Stake value_percent'],
                row['Held by ASPAC priority firm_Y_or_N'],
                row['PE priority firm '],
                row['Held since'],
                row['Held for more than three years_Y_N_NA'],
                row['KPMG credentials_PE_Company_Both_N'],
                row['KPMG firm'],
                row['Engagement partner']
            )
            for row in result
        ],
        {
            'Num': 0,
            'Picks': 1,
            'Captured in the week': 2,
            'Date': 3,
            'Lead type': 4,
            'Target HS industry classification': 5,
            'Dominant sector_as per MM': 6,
            'Sectors_as per MM': 7,
            'Sub Sectors_as per MM': 8,
            'HS Target region': 9,
            'Dominant country': 10,
            'Countries': 11,
            'Next Steps': 12,
            'Target': 13,
            'Target_Chinese': 14,
            'Deal intelligence': 15,
            'Short BD': 16,
            'Bidders': 17,
            'Sellers_or_Vendors': 18,
            'Type of transaction': 19,
            'Intelligence type_as per MM': 20,
            'Topics_as per MM': 21,
            'Value_USDm': 22,
            'Value description': 23,
            'Intelligence size': 24,
            'Intelligence size bucket': 25,
            'Stake value_percent': 26,
            'Held by ASPAC priority firm_Y_or_N': 27,
            'PE priority firm ': 28,
            'Held since': 29,
            'Held for more than three years_Y_N_NA': 30,
            'KPMG credentials_PE_Company_Both_N': 31,
            'KPMG firm': 32,
            'Engagement partner': 33
        }
    )


    # Extract row values and header positions
    row_values, header_positions = data_tuple

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Append the header row to the worksheet based on the header positions
    header_row = [None] * len(header_positions)
    for header, position in header_positions.items():
        header_row[position] = header
    ws.append(header_row)

    # Make the header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Append the data rows to the worksheet
    for row_data in row_values:
        # Check if row_data has at least 2 elements before converting date to string
        if len(row_data) >= 2 and isinstance(row_data[1], datetime.date):
            row_data = list(row_data)  # Convert tuple to list to allow modification
            row_data[1] = row_data[1].strftime('%Y-%m-%d')  # Convert date to string

        # Replace None values with an empty string
        row_data = ['' if value is None else value for value in row_data]

        # Replace '\r' and '\n' characters with line breaks in strings
        row_data = [value.replace('\r', '\n') if isinstance(value, str) else value for value in row_data]

        # Append the preprocessed row_data to the worksheet
        ws.append(row_data)

    # Create an in-memory file-like object to save the workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set response headers for file download
    headers = {
        'Content-Disposition': f'attachment; filename={"Market Scan.xlsx"}',
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "POST, GET, OPTIONS",
    }
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return Response(
        content=output.read(),
        headers=headers,
        media_type=media_type
    )


@app.get('/get_marketscan')
async def get_marketscan():
    table_name = "Market_Scan"
    job = """
    SELECT
        `Num`,
        `Picks`,
        `Captured in the week`,
        `Date`,
        `Lead type`,
        `Target HS industry classification`,
        `Dominant sector_as per MM`,
        `Sectors_as per MM`,
        `Sub Sectors_as per MM`,
        `HS Target region`,
        `Dominant country`,
        `Countries`,
        `Next Steps`,
        `Target`,
        `Target_Chinese`,
        `Deal intelligence`,
        `Short BD`,
        `Bidders`,
        `Sellers_or_Vendors`,
        `Type of transaction`,
        `Intelligence type_as per MM`,
        `Topics_as per MM`,
        `Value_USDm`,
        `Value description`,
        `Intelligence size`,
        `Intelligence size bucket`,
        `Stake value_percent`,
        `Held by ASPAC priority firm_Y_or_N`,
        `PE priority firm `,
        `Held since`,
        `Held for more than three years_Y_N_NA`,
        `KPMG credentials_PE_Company_Both_N`,
        `KPMG firm`,
        `Engagement partner`,
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Num;
    """.format(table_name=table_name)

    result = client.query(job)

    return_json = []

    # Iterate through rows
    for row in result:
        row_dict = {}
        
        # Iterate through columns dynamically
        for column_name in row.keys():
            row_dict[column_name] = row[column_name]

        return_json.append(row_dict)

    return return_json

@app.post('/update_marketscan')
async def update_marketscan(request: Request):
    table_name = "Market_Scan"

    try:
        data = await request.json()

    except JSONDecodeError:
        return 'Invalid JSON data.'
    
    success = True  # Assume success initially

    try:
        for row_data in data:
            row_id = row_data.get("row_id")
            updated_column = row_data.get("column_name")
            updated_value = row_data.get("edited_value")

            # Construct the UPDATE query dynamically
            update_query = f"""
            UPDATE `testing-bigquery-vertexai.templates.{table_name}`
            SET `{updated_column}` = '{updated_value}'
            WHERE `Num` = {row_id}
            """
            client.query(update_query)
            
    except Exception as e:
        print(e)
        success = False
    
    return {"success": success}

class MarketScanRow(BaseModel):
    Num: float
    Picks: bool
    Captured_in_the_week: str
    Date: float
    Lead_type: str
    Target_HS_industry_classification: str
    Dominant_sector_as_per_MM: str
    Sectors_as_per_MM: str
    Sub_Sectors_as_per_MM: str
    HS_Target_region: str
    Dominant_country: str
    Countries: str
    Next_Steps: str
    Target: str
    Target_Chinese: str
    Deal_intelligence: str
    Short_BD: str
    Bidders: str
    Sellers_or_Vendors: str
    Type_of_transaction: str
    Intelligence_type_as_per_MM: str
    Topics_as_per_MM: str
    Value_USDm: str
    Value_description: str
    Intelligence_size: str
    Intelligence_size_bucket: str
    Stake_value_percent: str
    Held_by_ASPAC_priority_firm_Y_or_N: str
    PE_priority_firm: str
    Held_since: str
    Held_for_more_than_three_years_Y_N_NA: str
    KPMG_credentials_PE_Company_Both_N: str
    KPMG_firm: str
    Engagement_partner: str

@app.post('/add_marketscan_row')
async def add_marketscan_row(request: Request, row_data: MarketScanRow):
    table_name = "Market_Scan"

    try:
        # Initialize the BigQuery client
        client = bigquery.Client()

        # Construct the INSERT query dynamically
        insert_query = f"""
        INSERT INTO `testing-bigquery-vertexai.templates.{table_name}`
        (Num, Picks, `Captured in the week`, Date, `Lead type`, `Target HS industry classification`,
        `Dominant sector_as per MM`, `Sectors_as per MM`, `Sub Sectors_as per MM`, `HS Target region`,
        `Dominant country`, Countries, `Next Steps`, Target, Target_Chinese, `Deal intelligence`,
        `Short BD`, Bidders, Sellers_or_Vendors, `Type of transaction`, `Intelligence type_as per MM`,
        `Topics_as per MM`, `Value_USDm`, `Value description`, `Intelligence size`, `Intelligence size bucket`,
        `Stake value_percent`, `Held by ASPAC priority firm_Y_or_N`, `PE priority firm `, `Held since`,
        `Held for more than three years_Y_N_NA`, `KPMG credentials_PE_Company_Both_N`, `KPMG firm`,
        `Engagement partner`)
        VALUES 
        ({row_data.Num}, {row_data.Picks}, '{row_data.Captured_in_the_week}', {row_data.Date}, 
        '{row_data.Lead_type}', '{row_data.Target_HS_industry_classification}', '{row_data.Dominant_sector_as_per_MM}',
        '{row_data.Sectors_as_per_MM}', '{row_data.Sub_Sectors_as_per_MM}', '{row_data.HS_Target_region}',
        '{row_data.Dominant_country}', '{row_data.Countries}', '{row_data.Next_Steps}', '{row_data.Target}',
        '{row_data.Target_Chinese}', '{row_data.Deal_intelligence}', '{row_data.Short_BD}', '{row_data.Bidders}',
        '{row_data.Sellers_or_Vendors}', '{row_data.Type_of_transaction}', '{row_data.Intelligence_type_as_per_MM}',
        '{row_data.Topics_as_per_MM}', '{row_data.Value_USDm}', '{row_data.Value_description}', 
        '{row_data.Intelligence_size}', '{row_data.Intelligence_size_bucket}', '{row_data.Stake_value_percent}',
        '{row_data.Held_by_ASPAC_priority_firm_Y_or_N}', '{row_data.PE_priority_firm}', '{row_data.Held_since}',
        '{row_data.Held_for_more_than_three_years_Y_N_NA}', '{row_data.KPMG_credentials_PE_Company_Both_N}',
        '{row_data.KPMG_firm}', '{row_data.Engagement_partner}')
        """

        # Run the query
        query_job = client.query(insert_query)

        # Wait for the query to complete (optional)
        query_job.result()

        return {"success": True}

    except Exception as e:
        print(e)
        return {"success": False}


@app.get("/download_rollingshortlist")
def download_RS():
    table_name = "Rolling_Shortlist"
    query = """
    SELECT
        `Num`,
        `Captured date`,
        `BP comments`,
        `Partner_or_Director`,
        `Target country`,
        `Sub sector`,
        `Source`,
        ` Target`,
        `Business Description`,
        `Financials`,
        `Revenue_USD_M`,
        `EBITDA_USD_M`,
        `Valuation_USD_M`,
        `Other Info`,
        `Deal Intelligence info`,
        `News Date`,
        `KPMG View - Redacted`,
        `Credentials`,
        `HS contact`,
        `Investment date`,
        `Geographic region`,
        `Asset pack `,
        `Target Region`,
        `Shareholders`,
        `Lead type`,
        `Target HS industry classification `,
        `Stake for sale_percent`,
        `Value_USD_M`,
        `Value description `,
        `Website`,
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Num;
    """.format(table_name=table_name)

    # Run the BigQuery query
    job = client.query(query)

    # Fetch the query results
    result = job.result()

    # Extract data rows and header positions
    data_tuple = (
        [
            (
                row.Num,
                row['Captured date'],
                row['BP comments'],
                row['Partner_or_Director'],
                row['Target country'],
                row['Sub sector'],
                row['Source'],
                row[' Target'],
                row['Business Description'],
                row['Financials'],
                row['Revenue_USD_M'],
                row['EBITDA_USD_M'],
                row['Valuation_USD_M'],
                row['Other Info'],
                row['Deal Intelligence info'],
                row['News Date'],
                row['KPMG View - Redacted'],
                row['Credentials'],
                row['HS contact'],
                row['Investment date'],
                row['Geographic region'],
                row['Asset pack '],
                row['Target Region'],
                row['Shareholders'],
                row['Lead type'],
                row['Target HS industry classification '],
                row['Stake for sale_percent'],
                row['Value_USD_M'],
                row['Value description '],
                row['Website'],
            )
            for row in result
        ],
        {
            'Num': 0,
            'Captured date': 1,
            'BP comments': 2,
            'Partner_or_Director': 3,
            'Target country': 4,
            'Sub sector': 5,
            'Source': 6,
            ' Target': 7,
            'Business Description': 8,
            'Financials': 9,
            'Revenue_USD_M': 10,
            'EBITDA_USD_M': 11,
            'Valuation_USD_M': 12,
            'Other Info': 13,
            'Deal Intelligence info': 14,
            'News Date': 15,
            'KPMG View - Redacted': 16,
            'Credentials': 17,
            'HS contact': 18,
            'Investment date': 19,
            'Geographic region': 20,
            'Asset pack ': 21,
            'Target Region': 22,
            'Shareholders': 23,
            'Lead type': 24,
            'Target HS industry classification ': 25,
            'Stake for sale_percent': 26,
            'Value_USD_M': 27,
            'Value description ': 28,
            'Website': 29,
        }
    )



    # Extract row values and header positions
    row_values, header_positions = data_tuple

    # Create a new Workbook
    wb = Workbook()
    ws = wb.active

    # Append the header row to the worksheet based on the header positions
    header_row = [None] * len(header_positions)
    for header, position in header_positions.items():
        header_row[position] = header
    ws.append(header_row)

    # Make the header row bold
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Append the data rows to the worksheet
    for row_data in row_values:
        # Check if row_data has at least 2 elements before converting date to string
        if len(row_data) >= 2 and isinstance(row_data[1], datetime.date):
            row_data = list(row_data)  # Convert tuple to list to allow modification
            row_data[1] = row_data[1].strftime('%Y-%m-%d')  # Convert date to string

        # Replace None values with an empty string
        row_data = ['' if value is None else value for value in row_data]

        # Replace '\r' and '\n' characters with line breaks in strings
        row_data = [value.replace('\r', '\n') if isinstance(value, str) else value for value in row_data]

        # Append the preprocessed row_data to the worksheet
        ws.append(row_data)

    # Create an in-memory file-like object to save the workbook
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Set response headers for file download
    headers = {
        'Content-Disposition': f'attachment; filename={"Rolling Shortlist.xlsx"}',
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Headers": "*",
        "Access-Control-Allow-Methods": "POST, GET, OPTIONS",
    }
    media_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

    return Response(
        content=output.read(),
        headers=headers,
        media_type=media_type
    )


@app.get('/get_rollingshortlist')
async def get_rollingshortlist():
    table_name = "Rolling_Shortlist"
    job = """
    SELECT
        `Num`,
        `Captured date`,
        `BP comments`,
        `Partner_or_Director`,
        `Target country`,
        `Sub sector`,
        `Source`,
        ` Target`,
        `Business Description`,
        `Financials`,
        `Revenue_USD_M`,
        `EBITDA_USD_M`,
        `Valuation_USD_M`,
        `Other Info`,
        `Deal Intelligence info`,
        `News Date`,
        `KPMG View - Redacted`,
        `Credentials`,
        `HS contact`,
        `Investment date`,
        `Geographic region`,
        `Asset pack `,
        `Target Region`,
        `Shareholders`,
        `Lead type`,
        `Target HS industry classification `,
        `Stake for sale_percent`,
        `Value_USD_M`,
        `Value description `,
        `Website`,
    FROM `testing-bigquery-vertexai.templates.{table_name}`
    ORDER BY Num;
    """.format(table_name=table_name)

    result = client.query(job)

    return_json = []

    # Iterate through rows
    for row in result:
        row_dict = {}
        
        # Iterate through columns dynamically
        for column_name in row.keys():
            row_dict[column_name] = row[column_name]

        return_json.append(row_dict)

    return return_json

@app.post('/update_rollingshortlist')
async def update_rollingshortlist(request: Request):
    table_name = "Rolling_Shortlist"

    try:
        data = await request.json()

    except JSONDecodeError:
        return 'Invalid JSON data.'
    
    success = True  # Assume success initially

    try:
        for row_data in data:
            row_id = row_data.get("row_id")
            updated_column = row_data.get("column_name")
            updated_value = row_data.get("edited_value")

            # Construct the UPDATE query dynamically
            update_query = f"""
            UPDATE `testing-bigquery-vertexai.templates.{table_name}`
            SET `{updated_column}` = '{updated_value}'
            WHERE `Num` = {row_id}
            """
            client.query(update_query)
            
    except Exception as e:
        print(e)
        success = False
    
    return {"success": success}

class RollingShortlistRow(BaseModel):
    Num: float
    Captured_date: float
    BP_comments: str
    Partner_or_Director: str
    Target_country: str
    Sub_sector: str
    Source: str
    Target: str
    Business_Description: str
    Financials: str
    Revenue_USD_M: str
    EBITDA_USD_M: str
    Valuation_USD_M: str
    Other_Info: str
    Deal_Intelligence_info: str
    News_Date: float
    KPMG_View_Redacted: str
    Credentials: str
    HS_contact: str
    Investment_date: str
    Geographic_region: str
    Asset_pack: str
    Target_Region: str
    Shareholders: str
    Lead_type: str
    Target_HS_industry_classification: str
    Stake_for_sale_percent: str
    Value_USD_M: str
    Value_description: str
    Website: str

@app.post('/add_rollingshortlist_row')
async def add_rollingshortlist_row(request: Request, row_data: RollingShortlistRow):
    table_name = "Rolling_Shortlist"

    try:
        # Initialize the BigQuery client
        client = bigquery.Client()

        # Construct the INSERT query dynamically
        insert_query = f"""
        INSERT INTO `testing-bigquery-vertexai.templates.{table_name}`
        (Num, `Captured date`, `BP comments`, `Partner_or_Director`, `Target country`, `Sub sector`, `Source`, ` Target`, `Business Description`, `Financials`, `Revenue_USD_M`, `EBITDA_USD_M`, `Valuation_USD_M`, `Other Info`, `Deal Intelligence info`, `News Date`, `KPMG View - Redacted`, `Credentials`, `HS contact`, `Investment date`, `Geographic region`, `Asset pack `, `Target Region`, `Shareholders`, `Lead type`, `Target HS industry classification `, `Stake for sale_percent`, `Value_USD_M`, `Value description `, `Website`) VALUES 
        ({row_data.Num}, {row_data.Captured_date}, '{row_data.BP_comments}', '{row_data.Partner_or_Director}', 
        '{row_data.Target_country}', '{row_data.Sub_sector}', '{row_data.Source}', '{row_data.Target}', 
        '{row_data.Business_Description}', '{row_data.Financials}', '{row_data.Revenue_USD_M}', 
        '{row_data.EBITDA_USD_M}', '{row_data.Valuation_USD_M}', '{row_data.Other_Info}', 
        '{row_data.Deal_Intelligence_info}', {row_data.News_Date}, '{row_data.KPMG_View_Redacted}', 
        '{row_data.Credentials}', '{row_data.HS_contact}', '{row_data.Investment_date}', 
        '{row_data.Geographic_region}', '{row_data.Asset_pack}', '{row_data.Target_Region}', 
        '{row_data.Shareholders}', '{row_data.Lead_type}', '{row_data.Target_HS_industry_classification}', 
        '{row_data.Stake_for_sale_percent}', '{row_data.Value_USD_M}', '{row_data.Value_description}', 
        '{row_data.Website}')
        """

        # Run the query
        query_job = client.query(insert_query)

        # Wait for the query to complete (optional)
        query_job.result()

        return {"success": True}

    except Exception as e:
        print(e)
        return {"success": False}

if __name__ == '__main__':
    import uvicorn
    uvicorn.run("preview:app", host='127.0.0.1', port=5003, reload=True)